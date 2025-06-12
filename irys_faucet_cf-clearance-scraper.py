import requests
import time
import openpyxl
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# 配置参数
MAX_WORKERS = 10  # 并发处理的最大线程数
USE_SIMPLE_TURNSTILE = False  # 是否使用代理获取CF验证码，不用改为：False
TURNSTILE_API = "http://http://localhost:3000/cf-clearance-scraper"  # Turnstile API 地址
NSTPROXY_API = ""  # NST代理 API 地址 注册购买地址：https://app.nstproxy.com/register?i=C38Qg4

# 用于线程间同步对 Excel 的写操作
excel_lock = threading.Lock()


def get_proxy():
    try:
        resp = requests.get(NSTPROXY_API, timeout=15)
        resp.raise_for_status()
        proxies = resp.json()
        if isinstance(proxies, list) and proxies:
            proxy = proxies[0]
            # 构建代理URL（用于requests请求）
            proxy_url = f"http://{proxy['username']}:{proxy['password']}@{proxy['ip']}:{proxy['port']}"
            print(f"[Proxy] 获取到代理: {proxy_url}")
            # 返回两种格式的代理信息
            return {
                # 用于 get_cf_turnstile 提交给后端的格式
                "proxy_info": {
                    "host": proxy["ip"],
                    "port": int(proxy["port"]),
                    "username": proxy["username"],
                    "password": proxy["password"]
                },
                # 用于 requests 请求的格式
                "http": proxy_url,
                "https": proxy_url
            }
        else:
            print(f"[Proxy] 接口返回格式异常: {proxies}")
    except requests.RequestException as e:
        print(f"[Proxy] 获取代理失败: {e}")
    except ValueError as e:
        print(f"[Proxy] 端口转换失败: {e}")
    return None


def get_cf_turnstile(proxy: dict = None):
    # 构建基础请求数据
    json_data = {
        'url': 'https://irys.xyz/api/faucet',
        'siteKey': '0x4AAAAAAA6vnrvBCtS4FAl-',
        'mode': 'turnstile-min'
    }
    
    # 如果有代理配置，添加到请求数据中
    if proxy and "proxy_info" in proxy:
        json_data['proxy'] = proxy["proxy_info"]

    try:
        response = requests.post(
            TURNSTILE_API,
            headers={
                'Content-Type': 'application/json'
            },
            json=json_data,
            timeout=60
        )
        result = response.json()
        if 'token' not in result:
            print(f"[Turnstile] 响应中没有 token: {result}")
            return None
        token = result['token']
        print(f"[Turnstile] 获取到验证码token：{token[:8]}...")
        return token
    except Exception as e:
        print(f"[Turnstile] 请求失败：{e}")
        return None

        
def submit_faucet(token: str, wallet_address: str, proxy: dict):
    url = "https://irys.xyz/api/faucet"
    payload = {"captchaToken": token, "walletAddress": wallet_address}
    headers = {"Content-Type": "application/json"}

    try:
        if proxy:
            resp = requests.post(url, json=payload, headers=headers, proxies=proxy, timeout=30)
        else:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)

        resp.raise_for_status()
        result = resp.json()

        if result.get("success") is True:
            print(f"[Faucet] {wallet_address} – 领水成功: {result}")
            return True, result.get("message", "")
        else:
            print(f"[Faucet] {wallet_address} – 领水失败: {result}")
            return False, result.get("message", "未知错误")

    except requests.RequestException as e:
        error_result = {}
        if e.response is not None:
            try:
                error_result = e.response.json()
                # 如果是 JSON 响应，只取 message 字段
                if isinstance(error_result, dict):
                    error_message = error_result.get("message", str(error_result))
                else:
                    error_message = str(error_result)
            except Exception:
                error_message = str(e.response.text)
        else:
            error_message = str(e)
        print(f"[Faucet] {wallet_address} – 领水失败: {error_message}")
        return False, error_message


def worker(task, ws, address_col, status_col, message_col, file_path):
    row, address = task
    print(f"🚰 正在处理：{address}")

    # 1. 先获取代理（每个线程只获取一次）
    proxy = get_proxy()
    if not proxy:
        print(f"[Main] {address} – 获取代理失败，标记为 false")
        with excel_lock:
            ws.cell(row=row, column=status_col).value = "false"
            ws.cell(row=row, column=message_col).value = "获取 代理 失败"
            ws.parent.save(file_path)
        return

    # 2. 使用代理获取 Turnstile token
    token = get_cf_turnstile(proxy)
    if not token:
        print(f"[Main] {address} – 获取 token 失败，标记为 false")
        with excel_lock:
            ws.cell(row=row, column=status_col).value = "false"
            ws.cell(row=row, column=message_col).value = "获取 验证码token 失败"
            ws.parent.save(file_path)
        return

    # 3. 使用获取到的代理提交 Faucet
    success, message = submit_faucet(token, address, proxy)
    with excel_lock:
        ws.cell(row=row, column=status_col).value = "success" if success else "false"
        ws.cell(row=row, column=message_col).value = message
        ws.parent.save(file_path)

    # 4. 短暂延迟，降低并发请求过快的风险
    time.sleep(1)

def process_wallets_from_excel(file_path: str, max_workers: int = MAX_WORKERS):
    """
    从 Excel 里批量读取地址和状态（"address" / "faucet"），并且并发执行领取流程。
    """
    # 载入工作簿并获取活动 sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 找到 "address"、"faucet" 和 "message" 列
    address_col = None
    status_col = None
    message_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "address":
            address_col = col
        elif header == "faucet":
            status_col = col
        elif header == "message":
            message_col = col

    if not address_col or not status_col:
        print("❌ Excel 缺少 'address' 或 'faucet' 列")
        return

    # 如果没有 message 列，添加一列
    if not message_col:
        message_col = ws.max_column + 1
        ws.cell(row=1, column=message_col).value = "message"
        wb.save(file_path)

    # 收集所有需要处理的地址
    tasks = []
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row=row, column=address_col).value
        status = ws.cell(row=row, column=status_col).value
        # 只跳过 status 已经是 "success" 的行，其它（None或"false"）都要处理
        if address and status != "success":
            tasks.append((row, address))

    # 随机打乱顺序，避免顺序性规律
    random.shuffle(tasks)

    # 使用线程池并发处理
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for task in tasks:
            futures.append(executor.submit(worker, task, ws, address_col, status_col, message_col, file_path))
        # 等待所有任务完成
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"[ThreadPool] 某个线程内抛出异常：{e}")

    print("✅ 所有地址处理完成，结果已写入 Excel。")

if __name__ == "__main__":
    process_wallets_from_excel("wallet.xlsx", max_workers=MAX_WORKERS)
