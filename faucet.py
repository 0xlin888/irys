import requests
import time
import openpyxl
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# 用于线程间同步对 Excel 的写操作
excel_lock = threading.Lock()

def get_proxy():
    url = ""  #这里用注册购买动态IP https://app.nstproxy.com/register?i=C38Qg4，把API粘贴进来
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        proxies = resp.json()
        if isinstance(proxies, list) and proxies:
            proxy = proxies[0]
            ip = proxy["ip"]
            port = proxy["port"]
            username = proxy["username"]
            password = proxy["password"]
            proxy_url = f"http://{username}:{password}@{ip}:{port}"
            print(f"[Proxy] 获取到代理: {proxy_url}")
            return {"http": proxy_url, "https": proxy_url}
        else:
            print(f"[Proxy] 接口返回格式异常: {proxies}")
    except requests.RequestException as e:
        print(f"[Proxy] 获取代理失败: {e}")
    return None

def get_cf_turnstile(proxy: dict):
    api_url = "http://api.nocaptcha.io/api/wanda/cloudflare/universal"
    payload = {
        "href": "https://irys.xyz/faucet",
        "proxy": proxy["http"] if proxy else "",   # 把 proxy_url 放到 payload 里（有的解决器 API 需要在 payload 里传）
        "sitekey": "0x4AAAAAAA6vnrvBCtS4FAl-"
    }
    headers = {
        "User-Token": "",#这里用注册购买打码服务 https://www.nocaptcha.io/register?c=HCLVQl，把Token粘贴进来
        "Content-Type": "application/json"
    }

    retries = 3
    for attempt in range(1, retries + 1):
        try:
            if proxy:
                resp = requests.post(api_url, json=payload, headers=headers, proxies=proxy, timeout=30)
            else:
                resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
            resp.raise_for_status()
            result = resp.json()

            # 假设返回格式为 {"status":1, "data": {"token": "xxx"}}
            if isinstance(result, dict) and result.get("status") == 1:
                token = result["data"]["token"]
                print(f"[Turnstile] 成功获取 token：{token[:8]}...")  # 只展示前 8 位
                return token
            else:
                print(f"[Turnstile] 第 {attempt} 次尝试，status != 1，返回：{result}")
        except requests.exceptions.RequestException as e:
            print(f"[Turnstile] 第 {attempt} 次尝试请求异常：{e}")

        if attempt < retries:
            time.sleep(2)

    print("[Turnstile] 三次尝试后未能获取 token")
    return None

def submit_faucet(token: str, wallet_address: str, proxy: dict):
    url = "https://irys.xyz/api/faucet"
    payload = {"captchaToken": token, "walletAddress": wallet_address}
    headers = {"Content-Type": "application/json"}

    try:
        if proxy:
            resp = requests.post(url, json=payload, headers=headers, proxies=proxy, timeout=30)
        else:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)

        resp.raise_for_status()
        result = resp.json()

        if result.get("success") is True:
            print(f"[Faucet] {wallet_address} – 领取成功: {result}")
            return True
        else:
            print(f"[Faucet] {wallet_address} – 领取失败: {result}")
            return False

    except requests.RequestException as e:
        error_result = {}
        if e.response is not None:
            try:
                error_result = e.response.json()
            except Exception:
                error_result = e.response.text
        print(f"[Faucet] {wallet_address} – 领取失败: {error_result}")
        return False

def worker(task, ws, address_col, status_col, file_path):
    row, address = task
    print(f"🚰 正在处理：{address}")

    # 1. 获取代理
    proxy = get_proxy()
    if not proxy:
        print(f"[Main] {address} – 未获取到代理，直接标记为 false")
        with excel_lock:
            ws.cell(row=row, column=status_col).value = "false"
            ws.parent.save(file_path)
        return

    # 2. 用代理去请求 Turnstile
    token = get_cf_turnstile(proxy)
    if not token:
        print(f"[Main] {address} – 获取 token 失败，标记为 false")
        with excel_lock:
            ws.cell(row=row, column=status_col).value = "false"
            ws.parent.save(file_path)
        return

    # 3. 提交 Faucet
    success = submit_faucet(token, address, proxy)
    with excel_lock:
        ws.cell(row=row, column=status_col).value = "success" if success else "false"
        ws.parent.save(file_path)

    # 4. 短暂延迟，降低并发请求过快的风险
    time.sleep(1)

def process_wallets_from_excel(file_path: str, max_workers: int = 10):
    """
    从 Excel 里批量读取地址和状态（“address” / “faucet”），并且并发执行领取流程。
    """
    # 载入工作簿并获取活动 sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 找到 “address” 和 “faucet” 列
    address_col = None
    status_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "address":
            address_col = col
        elif header == "faucet":
            status_col = col

    if not address_col or not status_col:
        print("❌ Excel 缺少 'address' 或 'faucet' 列")
        return

    # 收集所有需要处理的地址
    tasks = []
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row=row, column=address_col).value
        status = ws.cell(row=row, column=status_col).value
        # 只跳过 status 已经是 "success" 的行，其它（None或"false"）都要处理
        if address and status != "success":
            tasks.append((row, address))

    # 随机打乱顺序，避免顺序性规律
    random.shuffle(tasks)

    # 使用线程池并发处理
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for task in tasks:
            futures.append(executor.submit(worker, task, ws, address_col, status_col, file_path))
        # 等待所有任务完成
        for future in as_completed(futures):
            # 简单地让它跑完即可，如果需要捕获异常可以在这里处理
            try:
                future.result()
            except Exception as e:
                print(f"[ThreadPool] 某个线程内抛出异常：{e}")

    print("✅ 所有地址处理完成，结果已写入 Excel。")

if __name__ == "__main__":
    process_wallets_from_excel("wallet.xlsx", max_workers=10)
